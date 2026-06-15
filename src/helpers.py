import csv
import re
import shutil
import requests
from pathlib import Path
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT_DIR / "data"
CONFIG_DIR = ROOT_DIR / "config"
RESUME_DIR = ROOT_DIR / "resume"
COVER_LETTERS_DIR = ROOT_DIR / "cover_letters"

DATA_DIR.mkdir(exist_ok=True)
CONFIG_DIR.mkdir(exist_ok=True)
RESUME_DIR.mkdir(exist_ok=True)
COVER_LETTERS_DIR.mkdir(exist_ok=True)

CONFIG_FILE = CONFIG_DIR / "config.yaml"
CONFIG_EXAMPLE_FILE = CONFIG_DIR / "config.example.yaml"
AUTO_FILL_PROFILE_FILE = ROOT_DIR / "autofill_profile.json"
APPLIED_JOBS_FILE = DATA_DIR / "applied_jobs.csv"
SEEN_JOBS_FILE = DATA_DIR / "seen_jobs.csv"
APPLIED_COMPANIES_FILE = DATA_DIR / "applied_companies.csv"
REVIEWED_JOBS_FILE = DATA_DIR / "reviewed_jobs.csv"
JOBS_FILE = DATA_DIR / "jobs.csv"
APPLICATION_PLAN_CSV = DATA_DIR / "application_plan.csv"
APPLICATION_PLAN_XLSX = DATA_DIR / "application_plan.xlsx"
RESUME_FILE = RESUME_DIR / "resume.txt"

ADZUNA_API_BASE = "https://api.adzuna.com/v1/api/jobs"

EXCLUDE_TITLE_WORDS = ["senior", "manager", "director", "lead", "principal"]
PREFER_TITLE_WORDS = ["coordinator", "assistant", "associate", "entry level"]
EXCLUDE_TECHNICAL_ROLES = [
    "software engineer", "data scientist", "developer", "scientist",
    "staff engineer", "principal engineer", "machine learning engineer",
    "full stack developer", "backend engineer", "frontend engineer",
    "devops engineer", "kubernetes", "cloud architecture",
]
EXPERIENCE_PATTERNS = ["0-2 years", "entry level", "junior", "0-1 year", "1-2 years", "no experience required"]

SKIP_APPLIED_PATTERNS = [
    "fmi", "j.e. dunn", "insight global", "commonspirit",
    "govcio", "cbre", "kharon", "subaru",
]

HARD_EXCLUDE_TERMS = [
    "top secret", "secret clearance", "security clearance required", "active clearance",
    "ts/sci", "sci clearance", "polygraph", "dod clearance", "government clearance",
    "space operations", "satellite operations", "aerospace operations", "mission operations",
    "military operations", "cryptologic", "linguist", "mandarin required", "bilingual required",
    "fluent required", "phd required", "pe license", "cpa required", "series 7",
    "nurse", "rn", "clinical", "physician", "medical license",
]
CLEARANCE_DEFENSE_TERMS = [
    "top secret", "secret clearance", "security clearance", "active clearance",
    "ts/sci", "sci clearance", "polygraph", "dod clearance", "government clearance",
    "space operations", "satellite operations", "aerospace operations", "mission operations",
    "military operations", "cryptologic",
]
LICENSE_REQUIRED_TERMS = [
    "phd required", "pe license", "cpa required", "series 7", "medical license",
    "licensure", "license required", "certification required",
]
ADVANCED_CERTIFICATE_TERMS = [
    "pmp", "project management professional", "cissp", "security+", "aws certified",
    "ccna", "ccnp", "cisa", "cism", "six sigma black belt", "six sigma green belt",
    "professional engineer", "licensed professional engineer", "licensed practical nurse",
]
ENGINEERING_DEGREE_TERMS = [
    "engineering degree", "computer science degree", "software engineering degree",
    "bs in engineering", "b.s. in engineering", "bachelors in engineering",
]
BAD_COORDINATOR_TERMS = [
    "sales", "staffing", "field", "dispatch", "healthcare", "space", "defense",
    "technician", "warehouse", "logistics", "transportation", "shipping", "fleet",
    "maintenance", "construction", "hotel", "hospitality",
]
LOW_FIT_ROLE_TERMS = [
    "cybersecurity", "security engineer", "devops", "cloud engineer", "network engineer",
    "systems engineer", "data scientist", "research scientist", "scientist",
    "financial advisor", "accountant", "controller", "tax", "insurance sales",
    "law enforcement", "police", "detective", "armed security", "firefighter",
    "paramedic", "mechanic", "electrician", "plumber", "welder",
    "laboratory", "pharmaceutical", "clinical trials", "hospital", "medical office",
]
SAFE_5PLUS_ROLES = [
    "operations analyst", "business analyst", "business systems analyst",
    "implementation specialist", "technical operations", "data operations",
    "product operations", "revops", "workflow automation",
]
SKILL_TERMS = [
    "excel", "google sheets", "data entry", "project coordination", "communication",
    "organization", "workflow", "scheduling", "customer service", "presentation",
    "team support", "administrative support", "event planning", "social media",
    "analysis", "reporting", "calendar management",
]
TITLE_TERMS = [
    "operations coordinator", "administrative coordinator", "project coordinator",
    "marketing coordinator", "business analyst", "customer success associate",
    "event coordinator", "music industry coordinator", "assistant", "specialist",
]
ANALYST_DUPLICATE_TITLES = {
    "analyst", "business analyst", "operations analyst",
    "business systems analyst", "systems analyst",
}
RESUME_KEYWORD_TERMS = SKILL_TERMS + TITLE_TERMS + ["remote", "entry-level", "junior", "manager", "liaison", "support"]
TARGET_KEYWORDS = sorted(set(
    TITLE_TERMS + SAFE_5PLUS_ROLES + [
        "operations", "automation", "workflow", "customer success",
        "business operations", "project coordination", "data operations",
        "technical operations", "product operations", "revenue operations",
        "revops", "sales operations", "program coordinator",
        "technical support engineer", "process improvement",
        "onboarding", "reporting", "dashboards", "analytics", "crm",
        "salesforce", "jira", "excel", "ai", "automation",
    ]
))
ADJACENT_SEARCH_KEYWORDS = [
    "implementation", "coordinator", "support analyst", "operations associate",
    "project coordinator", "systems coordinator", "workflow analyst",
    "product operations", "revops", "customer success operations",
    "technical project coordinator", "program coordinator", "technical support engineer",
    "revenue operations analyst", "sales operations analyst", "business systems analyst",
    "data operations analyst", "qa analyst", "process improvement analyst",
    "ai operations associate", "workflow automation specialist",
]
CUSTOMER_SUCCESS_TERMS = [
    "customer success", "client success",
    "customer success associate", "customer success specialist",
]
TECH_ADJACENT_TITLE_TERMS = [
    "technical", "automation", "systems", "data",
    "workflow", "implementation", "saas", "crm", "devops",
]
ANALYST_TERMS = [
    "analyst", "operations analyst", "business analyst",
    "systems analyst", "data analyst", "qa analyst",
]
CORPORATE_OPS_TERMS = ["coordinator", "associate", "administrative", "operations", "assistant"]
TECH_FIT_TERMS = ["python", "sql", "api", "automation", "crm", "saas", "workflow", "data", "jira", "azure devops", "salesforce"]
EXPERIENCE_SKIP_PATTERNS = [
    "5+ years", "6+ years", "7+ years", "8+ years", "10+ years",
    "5 years", "6 years", "7 years", "8 years", "10 years",
    "3+ years", "4+ years", "3 years", "4 years",
    "experience required", "specialized experience required",
]
EXPERIENCE_APPLY_PATTERNS = [
    "0-2 years", "0-1 year", "1-2 years", "entry level",
    "junior", "training provided", "no experience required",
]
EXPERIENCE_MAYBE_PATTERNS = [
    "2+ years", "2 years", "3 years", "4 years",
    "mid level", "associate",
]

NEW_ROLE_CATEGORIES = {
    "japan_apac": ["japan", "apac", "asia-pacific", "tokyo"],
    "international_ops": ["international", "global", "worldwide"],
    "localization": ["localization", "localization specialist", "localization coordinator"],
    "gaming_entertainment": ["gaming", "game", "esports"],
    "music_events": ["music", "event", "events", "creator"],
    "travel_tech": ["travel", "tourism", "travel operations"],
    "tech_adjacent": TITLE_TERMS,
    "corporate_ops": ["coordinator", "associate", "administrative", "operations", "assistant"],
}

def load_config(config_file=CONFIG_FILE):
    defaults = {
        "daily_target_results": 20,
        "min_salary": 70000,
        "ideal_salary": 70000,
        "max_experience_years": 2,
        "allow_seen_jobs": False,
        "keywords": TARGET_KEYWORDS,
        "target_roles": TARGET_KEYWORDS,
        "exclude_keywords": [
            "senior", "director", "manager", "commission only",
            "door to door", "insurance sales", "independent contractor",
            "unpaid", "rn", "nurse", "clinical", "physician",
            "medical license", "legal experience required", "investment banking experience required",
        ],
        "high_priority_keywords": [],
        "results_per_search": 50,
        "results_pages": 3,
        "locations": ["Denver, CO", "remote"],
        "preferences": {
            "min_score": 50,
            "max_results": 50,
        },
    }

    try:
        with open(config_file, "r", encoding="utf-8") as f:
            loaded = __import__("yaml").safe_load(f) or {}
    except FileNotFoundError:
        loaded = {}
    except Exception:
        loaded = {}

    config = {**defaults, **loaded}
    if loaded.get("preferences"):
        config["preferences"] = {**defaults["preferences"], **loaded.get("preferences", {})}

    config["keywords"] = [kw for kw in (config.get("keywords") or []) if kw]
    config["target_roles"] = [role for role in (config.get("target_roles") or []) if role]
    config["exclude_keywords"] = [kw.lower() for kw in (config.get("exclude_keywords") or []) if kw]
    config["high_priority_keywords"] = [kw.lower() for kw in (config.get("high_priority_keywords") or []) if kw]
    return config


def fetch_real_jobs_adzuna(app_id, app_key, country="us", locations=None, results_per_search=20, pages=3, keywords=None):
    if locations is None:
        locations = ["Denver, CO", "remote"]
    if keywords is None:
        config = load_config()
        keywords = config.get("keywords", [])
    jobs = []
    for location in locations:
        for keyword in keywords:
            for page in range(1, pages + 1):
                try:
                    url = f"{ADZUNA_API_BASE}/{country}/search/{page}"
                    params = {
                        "app_id": app_id,
                        "app_key": app_key,
                        "what": keyword,
                        "where": location,
                        "results_per_page": results_per_search,
                        "content-type": "application/json",
                    }
                    response = requests.get(url, params=params, timeout=12)
                    if response.status_code != 200:
                        continue
                    data = response.json()
                    for result in data.get("results", []):
                        jobs.append({
                            "company": result.get("company", {}).get("display_name", ""),
                            "title": result.get("title", ""),
                            "location": result.get("location", {}).get("display_name", ""),
                            "link": result.get("redirect_url", ""),
                            "description": result.get("description", ""),
                            "created": result.get("created", ""),
                            "salary_min": result.get("salary_min", ""),
                            "salary_max": result.get("salary_max", ""),
                        })
                except Exception:
                    continue
    unique_jobs = []
    seen = set()
    for job in jobs:
        key = (normalize_match_key(job.get("company", "")), normalize_match_key(job.get("title", "")))
        if key not in seen:
            seen.add(key)
            unique_jobs.append(job)
    return unique_jobs


def fetch_real_jobs_fallback():
    return [
        {
            "job_id": "fallback-1",
            "title": "Operations Coordinator",
            "company": "Tech Startup Denver",
            "location": "Denver, CO",
            "link": "",
            "description": "Entry-level operations coordinator role supporting workflow, scheduling, and Excel reporting.",
            "salary_min": "55000",
            "salary_max": "65000",
        },
        {
            "job_id": "fallback-2",
            "title": "Business Operations Assistant",
            "company": "Marketing Firm",
            "location": "Remote",
            "link": "",
            "description": "Junior business operations assistant role with emphasis on coordination, CRM, and team support.",
            "salary_min": "60000",
            "salary_max": "70000",
        },
        {
            "job_id": "fallback-3",
            "title": "Administrative Coordinator",
            "company": "Denver Business Services",
            "location": "Denver, CO",
            "link": "",
            "description": "Administrative coordinator role with calendar management, reporting, and process improvement tasks.",
            "salary_min": "58000",
            "salary_max": "65000",
        },
    ]


def normalize_text(text):
    text = text or ""
    text = text.lower()
    text = re.sub(r"[^a-z0-9\s]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_resume_data(resume_file=RESUME_FILE):
    try:
        with open(resume_file, "r", encoding="utf-8") as f:
            text = f.read()
    except FileNotFoundError:
        return {"keywords": [], "skills": [], "titles": [], "full_text": ""}

    normalized = normalize_text(text)
    found_skills = [skill for skill in SKILL_TERMS if skill in normalized]
    found_titles = [title for title in TITLE_TERMS if title in normalized]
    found_keywords = [keyword for keyword in RESUME_KEYWORD_TERMS if keyword in normalized]

    keywords = list(dict.fromkeys(found_keywords))[:12] or ["operations", "administrative", "coordination"]
    skills = list(dict.fromkeys(found_skills))[:8] or ["communication", "organization", "team support"]
    titles = list(dict.fromkeys(found_titles))[:6] or ["operations coordinator", "assistant", "coordinator"]

    return {"keywords": keywords, "skills": skills, "titles": titles, "full_text": normalized}


def normalize_match_key(value):
    if not value:
        return ""
    text = value.lower()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\b(c/o|via|for client|for-client)\b", " ", text)
    text = re.sub(r"[^\w\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_title_variant(title):
    if not title:
        return ""
    normalized = title.lower()
    normalized = re.sub(r"\([^)]*\)", " ", normalized)
    normalized = re.sub(r"\b(c/o|via|for client|for-client)\b", " ", normalized)
    normalized = re.sub(r"[^\w\s]", " ", normalized)
    remove_tokens = [
        "remote", "hybrid", "onsite", "on site", "on-site",
        "entry", "level", "part time", "full time",
        "temporary", "contract", "ii", "iii", "iv", "v", "i", "sr", "jr",
    ]
    for token in remove_tokens:
        normalized = re.sub(r"\b" + re.escape(token) + r"\b", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def titles_similar(title1, title2):
    if not title1 or not title2:
        return False
    t1 = normalize_title_variant(title1)
    t2 = normalize_title_variant(title2)
    if not t1 or not t2:
        return False
    if t1 == t2:
        return True
    if t1 in ANALYST_DUPLICATE_TITLES and t2 in ANALYST_DUPLICATE_TITLES:
        return True
    if SequenceMatcher(None, t1, t2).ratio() > 0.90:
        return True
    words1 = set(t1.split())
    words2 = set(t2.split())
    if words1 and words2:
        shared = words1 & words2
        if len(shared) / min(len(words1), len(words2)) >= 0.75:
            return True
    return False


def entry_matches_job(entry, company_norm, title_norm, link_norm, exact_title=False):
    if entry.get("link") and link_norm and entry["link"] == link_norm:
        return True
    if entry.get("company") and entry["company"] == company_norm:
        if exact_title:
            return entry.get("title") == title_norm
        if entry.get("title") and titles_similar(entry["title"], title_norm):
            return True
        if entry.get("title") and entry["title"] in title_norm:
            return True
    return False


def generate_job_id(company, title):
    if not company or not title:
        return ""
    company_norm = re.sub(r"[^\w\s]", " ", company.lower())
    title_norm = re.sub(r"[^\w\s]", " ", title.lower())
    for word in ["remote", "hybrid", "i", "ii", "iii", "iv", "v", "senior", "junior", "entry", "level"]:
        company_norm = re.sub(r"\b" + re.escape(word) + r"\b", " ", company_norm)
        title_norm = re.sub(r"\b" + re.escape(word) + r"\b", " ", title_norm)
    company_norm = re.sub(r"\s+", " ", company_norm).strip()
    title_norm = re.sub(r"\s+", " ", title_norm).strip()
    return f"{company_norm}_{title_norm}".replace(" ", "_")


def normalize_link(value):
    if not value:
        return ""
    value = value.strip().lower()
    value = value.split("#", 1)[0]
    if "?" in value:
        base, query = value.split("?", 1)
        filtered = []
        for pair in query.split("&"):
            if "=" not in pair:
                continue
            name, val = pair.split("=", 1)
            if name.lower().startswith(("utm_", "fbclid", "gclid", "mc_")):
                continue
            filtered.append(f"{name}={val}")
        value = base
        if filtered:
            value = f"{base}?{'&'.join(filtered)}"
    return re.sub(r"/c/o/|/via/|/for-client/", "/", value)


def load_csv_history(path, fieldnames=None, pair_columns=("company", "title", "link")):
    if not Path(path).exists():
        return []
    entries = []
    try:
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                company_norm = normalize_match_key(row.get("company", ""))
                title_norm = normalize_match_key(row.get("title", ""))
                link_norm = normalize_link(row.get("link", ""))
                job_id = (row.get("job_id") or "").strip()
                reviewed_at = (row.get("reviewed_at") or "").strip()
                if company_norm or title_norm or link_norm or job_id:
                    entry = {
                        "company": company_norm,
                        "title": title_norm,
                        "link": link_norm,
                    }
                    if job_id:
                        entry["job_id"] = job_id
                    if reviewed_at:
                        entry["reviewed_at"] = reviewed_at
                    entries.append(entry)
    except Exception:
        return []
    return entries


def load_applied_jobs(applied_file=APPLIED_JOBS_FILE):
    return load_csv_history(applied_file)


def load_applied_from_plan(plan_file=APPLICATION_PLAN_CSV):
    if not Path(plan_file).exists():
        return []
    applied = []
    try:
        with open(plan_file, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                status = (row.get("status") or "").lower().strip()
                if status in ["applied", "skipped"]:
                    company_norm = normalize_match_key(row.get("company", ""))
                    title_norm = normalize_match_key(row.get("title", ""))
                    link_norm = normalize_link(row.get("link", ""))
                    if company_norm or title_norm or link_norm:
                        applied.append({"company": company_norm, "title": title_norm, "link": link_norm})
    except Exception:
        return []
    return applied


def load_seen_jobs(seen_file=SEEN_JOBS_FILE):
    return load_csv_history(seen_file)


def load_reviewed_jobs(reviewed_file=REVIEWED_JOBS_FILE):
    return load_csv_history(reviewed_file)


def expand_search_keywords(config):
    keywords = [kw for kw in (config.get("keywords") or []) if kw]
    for keyword in ADJACENT_SEARCH_KEYWORDS:
        if keyword not in keywords:
            keywords.append(keyword)
    return keywords


def load_applied_companies(applied_companies_file=APPLIED_COMPANIES_FILE):
    applied_companies = {}
    path = Path(applied_companies_file)
    if not path.exists():
        return applied_companies
    try:
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                company = row.get("company", "").strip()
                title = row.get("title", "").strip()
                if not company:
                    continue
                company_norm = normalize_match_key(company)
                applied_companies.setdefault(company_norm, set())
                if title:
                    applied_companies[company_norm].add(normalize_match_key(title))
    except Exception:
        pass
    return applied_companies


def titles_substantially_different(title1, title2):
    return not titles_similar(title1, title2)


def is_applied_or_excluded(job, applied_entries, applied_companies):
    job_company = (job.get("company") or "").lower()
    job_title = (job.get("title") or "").lower()
    job_link = normalize_link(job.get("link", ""))
    if not job_company and not job_title and not job_link:
        return False
    if any(block in job_company or block in job_title for block in SKIP_APPLIED_PATTERNS):
        return True
    job_company_norm = normalize_match_key(job_company)
    job_title_norm = normalize_match_key(job_title)
    if job_company_norm in applied_companies:
        applied_titles = applied_companies[job_company_norm]
        if job_title_norm and job_title_norm in applied_titles:
            return True
    for applied in applied_entries:
        if applied.get("link") and job_link and applied["link"] == job_link:
            return True
        if applied.get("company") == job_company_norm and applied.get("title") == job_title_norm:
            return True
    return False


def filter_applied_jobs(jobs, applied_entries, applied_companies):
    filtered = []
    applied_skipped = 0
    company_skipped = 0
    for job in jobs:
        if is_applied_or_excluded(job, applied_entries, applied_companies):
            job_company_norm = normalize_match_key(job.get("company", ""))
            if job_company_norm in applied_companies:
                company_skipped += 1
            else:
                applied_skipped += 1
            continue
        filtered.append(job)
    return filtered, applied_skipped + company_skipped, company_skipped


def is_seen_or_excluded(job, seen_entries, allow_seen=False):
    if allow_seen:
        return False
    job_id = (job.get("job_id") or "").strip()
    job_link = normalize_link(job.get("link", ""))
    if not job_id and not job_link:
        return False
    for seen in seen_entries:
        if seen.get("job_id") and job_id and seen["job_id"] == job_id:
            return True
        if seen.get("link") and job_link and seen["link"] == job_link:
            return True
    return False


def filter_seen_jobs(jobs, seen_entries, allow_seen=False):
    filtered = []
    skipped_count = 0
    for job in jobs:
        if is_seen_or_excluded(job, seen_entries, allow_seen=allow_seen):
            skipped_count += 1
            continue
        filtered.append(job)
    return filtered, skipped_count


def is_reviewed_or_excluded(job, reviewed_entries):
    job_company = (job.get("company") or "").lower()
    job_title = (job.get("title") or "").lower()
    job_link = normalize_link(job.get("link", ""))
    job_id = (job.get("job_id") or "").strip()
    if not job_company and not job_title and not job_link and not job_id:
        return False
    job_company_norm = normalize_match_key(job_company)
    job_title_norm = normalize_match_key(job_title)
    now = datetime.now()
    for reviewed in reviewed_entries:
        reviewed_at = reviewed.get("reviewed_at")
        if reviewed_at:
            try:
                reviewed_date = datetime.fromisoformat(reviewed_at)
            except ValueError:
                reviewed_date = None
            if reviewed_date and (now - reviewed_date).days >= 14:
                continue
        if reviewed.get("job_id") and job_id and reviewed["job_id"] == job_id:
            return True
        if reviewed.get("link") and job_link and reviewed["link"] == job_link:
            return True
        if reviewed.get("company") == job_company_norm and reviewed.get("title") == job_title_norm:
            return True
    return False


def filter_reviewed_jobs(jobs, reviewed_entries):
    filtered = []
    skipped_count = 0
    for job in jobs:
        if is_reviewed_or_excluded(job, reviewed_entries):
            skipped_count += 1
            continue
        filtered.append(job)
    return filtered, skipped_count


def filter_jobs(jobs, config=None):
    filtered = []
    stats = {"clearance_defense": 0, "licensed_healthcare": 0, "senior_experience": 0}
    keywords = [kw.lower() for kw in (config.get("keywords") or [])] if config else []
    exclude_keywords = [kw.lower() for kw in (config.get("exclude_keywords") or [])] if config else []
    min_salary = config.get("min_salary") if config else None
    include_healthcare = config.get("include_healthcare", False) if config else False
    for job in jobs:
        title = (job.get("title") or "").lower()
        description = (job.get("description") or "").lower()
        salary_min = job.get("salary_min")
        text = f"{title} {description}"
        healthcare_terms = ["clinical", "nurse", "rn", "prn", "physician", "medical", "healthcare"]
        if not include_healthcare and any(term in text for term in healthcare_terms):
            stats["licensed_healthcare"] += 1
            continue
        if any(term in text for term in CLEARANCE_DEFENSE_TERMS):
            stats["clearance_defense"] += 1
            continue
        if any(term in text for term in LICENSE_REQUIRED_TERMS):
            stats["licensed_healthcare"] += 1
            continue
        if any(term in text for term in HARD_EXCLUDE_TERMS):
            stats["licensed_healthcare"] += 1
            continue
        if any(term in title for term in ["coordinator", "service", "specialist"]):
            if any(bad in text for bad in BAD_COORDINATOR_TERMS):
                stats["licensed_healthcare"] += 1
                continue
        if any(pattern in text for pattern in ["3+ years", "3 years", "4+ years", "4 years", "5+ years", "5 years", "6+ years", "6 years", "7+ years", "7 years", "8+ years", "8 years", "9+ years", "9 years", "10+ years", "10 years", "experience required", "specialized experience required"]):
            if not any(safe in text for safe in SAFE_5PLUS_ROLES):
                stats["senior_experience"] += 1
                continue
        if any(keyword in text for keyword in exclude_keywords):
            continue
        if any(word in title for word in EXCLUDE_TITLE_WORDS):
            continue
        if any(role in title for role in EXCLUDE_TECHNICAL_ROLES):
            continue
        if min_salary and salary_min:
            try:
                if float(salary_min) < min_salary:
                    continue
            except (TypeError, ValueError):
                pass
        if any(term in title for term in LOW_FIT_ROLE_TERMS) or any(term in text for term in LOW_FIT_ROLE_TERMS):
            stats["licensed_healthcare"] += 1
            continue
        if "engineer" in title and "technical support engineer" not in title:
            if not any(term in title for term in ["service", "support", "operations", "systems"]):
                continue
        has_keyword_match = any(keyword in text for keyword in keywords) if keywords else False
        has_preferred = any(word in title for word in PREFER_TITLE_WORDS)
        has_entry_exp = any(pattern in description for pattern in EXPERIENCE_PATTERNS)
        has_target_text = any(term in text for term in [
            "operations", "business", "implementation", "workflow", "automation",
            "process improvement", "product operations", "revops", "customer success",
            "sales operations", "revenue operations", "data operations", "technical support",
            "onboarding", "reporting", "dashboards", "analytics", "crm", "salesforce",
            "jira", "excel", "cross-functional", "coordination", "project support",
            "systems",
        ])
        if not (has_keyword_match or has_preferred or has_entry_exp or has_target_text):
            continue
        job["job_id"] = generate_job_id(job.get("company"), job.get("title"))
        filtered.append(job)
    return filtered, stats


def filter_duplicates(jobs, title_similarity_threshold=0.90):
    seen = set()
    filtered = []
    skipped_count = 0
    titles_by_company = {}
    for job in jobs:
        link = normalize_link(job.get("link", ""))
        company_norm = normalize_match_key(job.get("company", ""))
        title_variant = normalize_title_variant(job.get("title", ""))
        key = ("link", link) if link else ("company_title", company_norm, title_variant)
        duplicate = False
        if key in seen:
            duplicate = True
        else:
            for existing_title in titles_by_company.get(company_norm, []):
                if existing_title and title_variant:
                    if existing_title == title_variant:
                        duplicate = True
                        break
                    if existing_title in ANALYST_DUPLICATE_TITLES and title_variant in ANALYST_DUPLICATE_TITLES:
                        duplicate = True
                        break
                    if SequenceMatcher(None, existing_title, title_variant).ratio() > title_similarity_threshold:
                        duplicate = True
                        break
        if duplicate:
            skipped_count += 1
            continue
        seen.add(key)
        filtered.append(job)
        if company_norm:
            titles_by_company.setdefault(company_norm, []).append(title_variant)
    return filtered, skipped_count


def ensure_cover_letters_folder():
    folder = COVER_LETTERS_DIR / datetime.now().strftime("%Y-%m-%d")
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def normalize_pdf_text(text):
    if not text:
        return ""
    replacements = {"’": "'", "‘": "'", "“": '"', "”": '"', "–": '-', "—": '-', "•": '-', "…": '...'}
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = text.replace("\t", " ")
    return re.sub(r"[ ]{2,}", " ", text).strip()


def find_existing_cover_letter(filename):
    for path in COVER_LETTERS_DIR.rglob(filename):
        if path.is_file():
            return path
    return None


def archive_current_application_plan(application_plan_xlsx=APPLICATION_PLAN_XLSX):
    source = Path(application_plan_xlsx)
    if not source.exists():
        return None
    archive_folder = DATA_DIR / "archive" / datetime.now().strftime("%Y-%m-%d")
    archive_folder.mkdir(parents=True, exist_ok=True)
    destination = archive_folder / source.name
    if destination.exists():
        destination = archive_folder / f"{source.stem}_{datetime.now().strftime('%H%M%S')}{source.suffix}"
    try:
        shutil.move(str(source), str(destination))
        return str(destination)
    except Exception:
        return None


def select_pdf_font(pdf):
    fonts = [
        ("Inter", "/Library/Fonts/Inter-Regular.ttf"),
        ("Calibri", "/Library/Fonts/Calibri.ttf"),
        ("Aptos", "/Library/Fonts/Aptos-Regular.ttf"),
    ]
    for name, path in fonts:
        if Path(path).exists():
            try:
                pdf.add_font(name, "", path, uni=True)
                return name
            except Exception:
                continue
    return "Helvetica"


def sanitize_filename(name, max_length=40):
    name = re.sub(r"[^A-Za-z0-9_ -]", "", name)
    name = re.sub(r"\s+", "_", name).strip("_-")
    return name[:max_length].rstrip("_-")


def generate_cover_letter_pdf(job, resume_data, cover_letters_folder=None):
    if cover_letters_folder is None:
        cover_letters_folder = ensure_cover_letters_folder()
    company = sanitize_filename(job.get("company", "Company")) or "Company"
    role = sanitize_filename(job.get("title", "Role")) or "Role"
    filename = f"{company}_{role}_CL.pdf"
    existing = find_existing_cover_letter(filename)
    if existing:
        return str(existing)
    filepath = cover_letters_folder / filename
    if filepath.exists():
        return str(filepath)
    pdf = FPDF(format="letter", unit="mm")
    pdf.set_auto_page_break(True, margin=18)
    pdf.add_page()
    font_name = select_pdf_font(pdf)
    pdf.set_font(font_name, size=12)
    pdf.set_left_margin(18)
    pdf.set_right_margin(18)
    pdf.set_y(20)
    pdf.set_font(font_name, "B", 16)
    header_text = normalize_pdf_text(f"{job.get('company','Hiring Team')} - {job.get('title','')}")
    pdf.cell(0, 8, header_text[:80], ln=1)
    pdf.ln(3)
    pdf.set_font(font_name, size=11)
    opening = f"Dear {job.get('company','Hiring Team')},"
    pdf.multi_cell(0, 6, normalize_pdf_text(opening))
    pdf.ln(3)
    skills = resume_data.get("skills", [])
    skill_phrase = ", ".join(skills[:3]) if skills else "operations, Excel, and process improvement"
    paragraph_one = (
        f"I’m excited about this {job.get('title','role')} because it aligns with my background in operations, process improvement, and business systems. "
        f"I bring hands-on experience using {skill_phrase} to streamline workflows, manage data, and support strong team coordination."
    )
    pdf.multi_cell(0, 6, normalize_pdf_text(paragraph_one))
    pdf.ln(2)
    paragraph_two = (
        f"In past roles, I helped teams improve visibility, reduce manual work, and keep projects moving through better communication and reliable follow-through. "
        f"I’m confident I can add value quickly by supporting your operations, reporting, and business systems efforts for this position."
    )
    pdf.multi_cell(0, 6, normalize_pdf_text(paragraph_two))
    pdf.ln(2)
    paragraph_three = (
        "I’d welcome the chance to discuss how I can contribute to your team and help this role move forward. "
        "Thank you for your consideration."
    )
    pdf.multi_cell(0, 6, normalize_pdf_text(paragraph_three))
    pdf.ln(6)
    closing = "Sincerely,\n[Your Name]"
    pdf.multi_cell(0, 6, normalize_pdf_text(closing))
    try:
        pdf.output(str(filepath))
        return str(filepath)
    except Exception:
        return ""


def save_application_plan_csv(rows, output_file=APPLICATION_PLAN_CSV):
    fieldnames = [
        "company", "title", "link", "fit_score", "priority",
        "salary_min", "salary_max", "interview_probability",
        "role_category", "tech_fit_reason", "reason_to_apply",
        "realistic_fit_score", "stretch_level", "next_skill_to_learn",
    ]
    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})
    return output_file


def save_application_plan_xlsx(rows, xlsx_file=APPLICATION_PLAN_XLSX):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Application Plan"
    fieldnames = [
        "company", "title", "link", "fit_score", "priority",
        "salary_min", "salary_max", "interview_probability",
        "role_category", "tech_fit_reason", "reason_to_apply",
        "realistic_fit_score", "stretch_level", "next_skill_to_learn",
    ]
    header_fill = PatternFill(start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, title in enumerate(fieldnames, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_center
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(fieldnames, start=1):
            value = row.get(key, "")
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if key == "company":
                cell.font = Font(color="0000B0FF")
            if key == "link" and value:
                cell.font = Font(color="0000B0FF", underline="single")
            if key == "priority":
                priority = str(value).lower()
                if "high" in priority:
                    cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
                elif "medium" in priority:
                    cell.fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
                elif "low" in priority:
                    cell.fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
            if key == "stretch_level" and str(value).lower() == "high":
                cell.fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    for col_idx in range(1, len(fieldnames) + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 18
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}1"
    sheet.freeze_panes = "A2"
    workbook.save(xlsx_file)
    return xlsx_file


def append_seen_jobs(jobs, seen_file=SEEN_JOBS_FILE):
    existing = load_seen_jobs(seen_file)
    seen_keys = {(entry.get("company"), entry.get("title"), entry.get("link")) for entry in existing}
    rows = []
    for job in jobs:
        key = (normalize_match_key(job.get("company", "")), normalize_match_key(job.get("title", "")), normalize_link(job.get("link", "")))
        if key in seen_keys:
            continue
        seen_keys.add(key)
        rows.append({
            "job_id": job.get("job_id", ""),
            "company": job.get("company", ""),
            "title": job.get("title", ""),
            "link": job.get("link", ""),
            "salary_min": job.get("salary_min", ""),
            "salary_max": job.get("salary_max", ""),
            "score": job.get("score", ""),
        })
    if not rows:
        return 0
    file_exists = Path(seen_file).exists()
    with open(seen_file, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["job_id", "company", "title", "link", "salary_min", "salary_max", "score"])
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def save_results(jobs, output_file=JOBS_FILE):
    if not jobs:
        return
    fieldnames = [
        "job_id", "company", "title", "location", "link", "fit_score", "priority",
        "match_reasons", "red_flags", "salary_min", "salary_max",
    ]
    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for job in jobs:
            writer.writerow({
                "job_id": job.get("job_id", ""),
                "company": job.get("company", ""),
                "title": job.get("title", ""),
                "location": job.get("location", ""),
                "link": job.get("link", ""),
                "fit_score": job.get("score", ""),
                "priority": job.get("priority", ""),
                "match_reasons": " | ".join(job.get("match_reasons", [])) if isinstance(job.get("match_reasons"), list) else job.get("match_reasons", ""),
                "red_flags": " | ".join(job.get("red_flags", [])) if isinstance(job.get("red_flags"), list) else job.get("red_flags", ""),
                "salary_min": job.get("salary_min", ""),
                "salary_max": job.get("salary_max", ""),
            })


def cleanup_generated_files(args):
    cleanup_type = "all_generated"
    older_than_days = None
    should_archive = False
    confirmed = False
    for arg in args:
        if arg.startswith("--type="):
            cleanup_type = arg.split("=", 1)[1]
        elif arg.startswith("--older-than="):
            try:
                older_than_days = int(arg.split("=", 1)[1])
            except ValueError:
                return
        elif arg == "--archive":
            should_archive = True
        elif arg == "--confirm":
            confirmed = True
    folders = []
    extensions = [".pdf", ".xlsx", ".csv"]
    if cleanup_type == "cover_letters":
        folders = [COVER_LETTERS_DIR]
    elif cleanup_type == "data":
        folders = [DATA_DIR]
    else:
        folders = [COVER_LETTERS_DIR, DATA_DIR]
    files_to_delete = []
    cutoff = None
    if older_than_days is not None:
        cutoff = datetime.now() - timedelta(days=older_than_days)
    for folder in folders:
        if not folder.exists():
            continue
        for file_path in folder.glob("**/*"):
            if not file_path.is_file():
                continue
            if file_path.suffix not in extensions:
                continue
            if file_path == APPLICATION_PLAN_CSV or file_path == APPLICATION_PLAN_XLSX:
                continue
            if file_path.name in {APPLIED_COMPANIES_FILE.name, SEEN_JOBS_FILE.name, REVIEWED_JOBS_FILE.name, CONFIG_FILE.name, CONFIG_EXAMPLE_FILE.name, RESUME_FILE.name}:
                continue
            if cutoff and datetime.fromtimestamp(file_path.stat().st_mtime) > cutoff:
                continue
            files_to_delete.append(file_path)
    if not files_to_delete:
        return []
    if not confirmed:
        return files_to_delete
    if should_archive:
        archive = ROOT_DIR / "backups" / datetime.now().strftime("%Y%m%d_%H%M%S")
        archive.mkdir(parents=True, exist_ok=True)
        for file_path in files_to_delete:
            shutil.copy2(file_path, archive / file_path.name)
        return files_to_delete
    for file_path in files_to_delete:
        try:
            file_path.unlink()
        except Exception:
            pass
    return files_to_delete


def mark_applied_companies(plan_file=APPLICATION_PLAN_CSV, applied_companies_file=APPLIED_COMPANIES_FILE):
    plan_path = Path(plan_file)
    if not plan_path.exists():
        return 0
    applied_companies = load_applied_companies(applied_companies_file)
    added = 0
    try:
        with plan_path.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                company = row.get("company", "").strip()
                title = row.get("title", "").strip()
                if not company:
                    continue
                company_norm = normalize_match_key(company)
                applied_companies.setdefault(company_norm, set())
                title_norm = normalize_match_key(title) if title else ""
                if title_norm and title_norm not in applied_companies[company_norm]:
                    applied_companies[company_norm].add(title_norm)
                    added += 1
    except Exception:
        return 0
    with open(applied_companies_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["company", "title"])
        for company_norm in sorted(applied_companies.keys()):
            for title_norm in sorted(applied_companies[company_norm]):
                writer.writerow([company_norm, title_norm])
    return added


def mark_reviewed_jobs(plan_file=APPLICATION_PLAN_CSV, reviewed_file=REVIEWED_JOBS_FILE):
    plan_path = Path(plan_file)
    if not plan_path.exists():
        return 0
    existing = load_reviewed_jobs(reviewed_file)
    reviewed_keys = {(entry.get("company"), entry.get("title"), entry.get("link")) for entry in existing}
    added = 0
    rows = []
    try:
        with plan_path.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                company = row.get("company", "").strip()
                title = row.get("title", "").strip()
                link = row.get("link", "").strip()
                key = (normalize_match_key(company), normalize_match_key(title), normalize_link(link))
                if not company or key in reviewed_keys:
                    continue
                reviewed_keys.add(key)
                rows.append({
                    "company": company,
                    "title": title,
                    "link": link,
                    "reviewed_at": datetime.now().isoformat(),
                })
                added += 1
    except Exception:
        return 0
    file_exists = Path(reviewed_file).exists()
    with open(reviewed_file, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["company", "title", "link", "reviewed_at"])
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)
    return added
